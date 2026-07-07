from pathlib import Path
from datetime import datetime
import shutil

from openpyxl import load_workbook

from app.models.training_session import TrainingSession


class WorkbookManager:

    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.filename = None

    def open(self, filename):
        """Open an Excel workbook."""

        self.filename = filename
        self.workbook = load_workbook(filename)

        # For now, use the active worksheet.
        # Later we'll let the user choose.
        self.worksheet = self.workbook.active

        # Return worksheet names so the UI can display them.
        return self.workbook.sheetnames

    def get_sheet_names(self):
        if self.workbook is None:
            return []

        return self.workbook.sheetnames

    def get_headers(self):
        if self.worksheet is None:
            return []

        return [cell.value for cell in self.worksheet[1]]

    def get_sessions(self):
        if self.worksheet is None:
            return []

        sessions = []

        for row in self.worksheet.iter_rows(min_row=2):

            values = [cell.value for cell in row]

            sessions.append(
                TrainingSession(
                    row_number=row[0].row,
                    values=values
                )
            )

        return sessions

    def save(self):
        if self.workbook is None:
            return

        self.backup()
        self.workbook.save(self.filename)

    def backup(self):
        if self.filename is None:
            return

        backup_folder = Path("backups")
        backup_folder.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        destination = (
            backup_folder /
            f"{timestamp}_{Path(self.filename).name}"
        )

        shutil.copy2(self.filename, destination)