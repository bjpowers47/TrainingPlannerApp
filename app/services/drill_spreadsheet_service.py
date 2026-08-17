"""Create and import the collaborative Development Library spreadsheet."""

from __future__ import annotations

from contextlib import closing
from dataclasses import dataclass, field
from html import unescape
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.drill import Drill
from app.models.player_development import (
    DEVELOPMENT_BLOCKS,
    get_block_by_name,
)
from app.services.coaching_library import (
    get_coaching_focus_by_id,
    get_coaching_focus_id_by_name,
)


HEADERS = (
    "Development Block", "Drill Name", "Purpose", "Duration Minutes",
    "Recommended Players", "Equipment", "Coaching Points", "Progressions",
    "Variations", "Notes",
)

EXPORT_HEADERS = (
    "Drill ID", "Development Block", "Coaching Focus", "Drill Name",
    "Purpose", "Duration Minutes", "Recommended Players",
    "Use Execution Details", "Sets", "Reps", "Work Minutes", "Rest Minutes",
    "Equipment", "Coaching Points", "Progressions", "Variations", "Notes",
)

LEGACY_EXPORT_HEADERS = (
    "Drill ID", "Development Block", "Coaching Focus", "Drill Name",
    "Purpose", "Duration Minutes", "Recommended Players",
    "Use Execution Details", "Sets", "Reps", "Work Seconds", "Rest Seconds",
    "Equipment", "Coaching Points", "Progressions", "Variations", "Notes",
)

# New templates use the complete round-trip schema. Drill ID is intentionally
# included for compatibility with exported workbooks, but may be left blank:
# imports always allocate an ID from the destination database.
TEMPLATE_HEADERS = EXPORT_HEADERS


def _safe_spreadsheet_text(value: object) -> object:
    """Prevent exported text from being interpreted as an Excel formula."""
    if not isinstance(value, str):
        return value
    if value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _pipe_join(values: list[str]) -> str:
    return " | ".join(str(value).strip() for value in values if str(value).strip())


def export_spreadsheet(filename: str | Path, repository, block_repository=None) -> int:
    """Export active drills to a formatted, coach-readable Excel workbook."""
    drills = [drill for drill in repository.get_all() if drill.active]
    blocks_by_id = (
        {block.id: block for block in block_repository.list_all()}
        if block_repository is not None
        else {block.id: block for block in DEVELOPMENT_BLOCKS}
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Drills"
    worksheet.append(EXPORT_HEADERS)

    for drill in drills:
        block = blocks_by_id.get(drill.development_block_id)
        legacy_focus = (
            get_coaching_focus_by_id(drill.technical_focus_id)
            if drill.technical_focus_id is not None
            else None
        )
        focus_name = drill.coaching_focus or (legacy_focus.name if legacy_focus else "")
        values = (
            drill.id,
            block.name if block else f"Unknown block ({drill.development_block_id})",
            focus_name,
            drill.name,
            drill.purpose,
            drill.duration_minutes,
            drill.recommended_players,
            "Yes" if drill.use_execution_details else "No",
            drill.sets,
            drill.reps,
            drill.work_seconds / 60 if drill.work_seconds is not None else None,
            drill.rest_seconds / 60 if drill.rest_seconds is not None else None,
            _pipe_join(drill.equipment),
            _pipe_join(drill.coaching_points),
            _pipe_join(drill.progressions),
            _pipe_join(drill.variations),
            drill.notes,
        )
        worksheet.append([_safe_spreadsheet_text(value) for value in values])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 24
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, column in enumerate(worksheet.columns, start=1):
        longest = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(longest + 2, 12), 42
        )

    workbook.save(filename)
    return len(drills)


@dataclass
class ImportReport:
    imported: int = 0
    duplicates: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def create_template(filename: str | Path) -> None:
    """Create an Excel template that coaches can complete without the app."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Drills"
    worksheet.append(TEMPLATE_HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(TEMPLATE_HEADERS))}1"
    worksheet.row_dimensions[1].height = 24
    for index, header in enumerate(TEMPLATE_HEADERS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(len(header) + 2, 12), 28
        )

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["How to use this template"])
    instructions.append(["Add one drill per row on the Drills sheet."])
    instructions.append(["Drill ID is optional; leave it blank when creating a new drill."])
    instructions.append(["Development Block and Drill Name are required."])
    instructions.append(["Use one of these Development Blocks:"])
    for block in DEVELOPMENT_BLOCKS:
        instructions.append([block.name])
    instructions.append(["Use Yes or No for Use Execution Details."])
    instructions.append(["Separate Equipment, Coaching Points, Progressions, and Variations with |."])
    workbook.save(filename)


def import_spreadsheet(filename: str | Path, repository, block_repository=None) -> ImportReport:
    """Validate a drill spreadsheet and add valid, non-duplicate drills."""
    report = ImportReport()
    workbook = load_workbook(filename, data_only=True)
    worksheet = next(
        (sheet for sheet in workbook.worksheets if sheet.title in {"Drills", "Active Drills"}),
        None,
    )
    if worksheet is None:
        report.errors.append('The spreadsheet must contain a "Drills" or "Active Drills" sheet.')
        return report
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    if tuple(headers) not in (HEADERS, EXPORT_HEADERS, LEGACY_EXPORT_HEADERS):
        report.errors.append("The Drills sheet does not use a supported column format.")
        return report
    is_export = tuple(headers) in (EXPORT_HEADERS, LEGACY_EXPORT_HEADERS)
    uses_minutes = tuple(headers) == EXPORT_HEADERS

    # Resolve foreign keys from the target database. IDs in an exported file or
    # in the built-in display library are not portable between databases.
    if hasattr(repository, "get_import_reference_data"):
        blocks, focuses = repository.get_import_reference_data()
    else:
        blocks = [{"id": block.id, "name": block.name} for block in DEVELOPMENT_BLOCKS]
        focuses = []

    def normalized(value: object) -> str:
        text = unescape(str(value or "")).casefold().replace("&", " and ")
        return re.sub(r"[^a-z0-9]+", " ", text).strip()

    block_aliases = {
        "1v1": "1v1 moves",
        "speed": "speed and agility",
    }
    block_by_name = {normalized(row["name"]): row for row in blocks}
    for alias, canonical in block_aliases.items():
        if canonical in block_by_name:
            block_by_name[alias] = block_by_name[canonical]

    existing = repository.get_all()
    known = {(drill.development_block_id, drill.name.casefold()) for drill in existing}
    next_id = max((drill.id for drill in existing), default=0) + 1

    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in values):
            continue
        if is_export:
            (_exported_id, block_name, focus_name, name, purpose, duration,
             players, use_details, sets, reps, work, rest, equipment, points,
             progressions, variations, notes) = values
        else:
            block_name, name, purpose, duration, players, equipment, points, progressions, variations, notes = values
            focus_name = None
            use_details = False
            sets = reps = work = rest = None
        block = block_by_name.get(normalized(block_name)) if block_name else None
        clean_name = str(name).strip() if name else ""
        if not clean_name:
            report.errors.append(f"Row {row_number}: Drill Name is required.")
            continue
        if block is None and block_name and block_repository is not None:
            created = block_repository.ensure_active(str(block_name))
            block = {"id": created.id, "name": created.name}
            block_by_name[normalized(created.name)] = block
        if block is None:
            supplied = str(block_name or "").strip() or "(blank)"
            report.errors.append(
                f"Row {row_number}: Development Block '{supplied}' is not active or configured."
            )
            continue
        block_id = block["id"]
        key = (block_id, clean_name.casefold())
        if key in known:
            report.duplicates.append(f"Row {row_number}: {clean_name}")
            continue
        try:
            duration_minutes = int(duration or 0)
        except (TypeError, ValueError):
            report.errors.append(f"Row {row_number}: Duration Minutes must be a whole number.")
            continue
        focus_id = None
        if focus_name and focuses:
            focus = next((row for row in focuses if
                row["development_block_id"] == block_id and
                normalized(row["name"]) == normalized(focus_name)), None)
            if focus is not None:
                focus_id = focus["id"]

        def whole_number(value, label):
            if value in (None, ""):
                return None
            try:
                return int(value)
            except (TypeError, ValueError):
                raise ValueError(label)

        try:
            execution_values = [whole_number(value, label) for value, label in (
                (sets, "Sets"), (reps, "Reps"),
            )]
            if uses_minutes:
                timed_values = []
                for value, label in ((work, "Work Minutes"), (rest, "Rest Minutes")):
                    if value in (None, ""):
                        timed_values.append(None)
                        continue
                    minutes = float(value)
                    if minutes < 0:
                        raise ValueError(label)
                    timed_values.append(round(minutes * 60))
            else:
                timed_values = [whole_number(value, label) for value, label in (
                    (work, "Work Seconds"), (rest, "Rest Seconds"),
                )]
        except ValueError as error:
            report.errors.append(f"Row {row_number}: {error} must be a valid non-negative number.")
            continue

        separator = "|" if is_export else ";"
        split = lambda value: [item.strip() for item in str(value or "").split(separator) if item.strip()]
        repository.save(Drill(
            id=next_id, name=clean_name, development_block_id=block_id,
            technical_focus_id=focus_id,
            coaching_focus=str(focus_name or "").strip(),
            purpose=str(purpose or "").strip(), duration_minutes=duration_minutes,
            recommended_players=str(players or "").strip(), equipment=split(equipment),
            coaching_points=split(points), progressions=split(progressions),
            variations=split(variations), notes=str(notes or "").strip(),
            use_execution_details=normalized(use_details) in {"yes", "true", "1"},
            sets=execution_values[0], reps=execution_values[1],
            work_seconds=timed_values[0], rest_seconds=timed_values[1],
        ))
        known.add(key)
        next_id += 1
        report.imported += 1
    return report
