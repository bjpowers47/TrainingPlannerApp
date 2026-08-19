import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.models.drill import Drill
from app.services.drill_spreadsheet_service import (
    EXPORT_HEADERS,
    TEMPLATE_HEADERS,
    create_template,
    export_spreadsheet,
    import_spreadsheet,
)


class _Repository:
    def __init__(self, drills=None):
        self.drills = drills

    def get_all(self):
        if self.drills is not None:
            return self.drills
        return [
            Drill(
                id=7,
                name="Passing Pattern",
                development_block_id=2,
                technical_focus_id=11,
                purpose="=unsafe formula text",
                duration_minutes=15,
                recommended_players="8–12",
                use_execution_details=True,
                sets=3,
                reps=4,
                work_seconds=45,
                rest_seconds=20,
                equipment=["Balls", "Cones"],
                coaching_points=["Scan", "Open body"],
                progressions=["One touch"],
                variations=["Add defender"],
                notes="Coach notes",
            ),
            Drill(id=8, name="Archived", development_block_id=1, active=False),
        ]

    def save(self, drill):
        self.drills.append(drill)


class _BlockRepository:
    def list_all(self):
        return [
            type("Block", (), {"id": 610, "name": "Receiving & Passing"})(),
        ]


class DrillSpreadsheetExportTests(unittest.TestCase):
    def test_exports_active_drills_with_labels_formatting_and_safe_text(self):
        with tempfile.TemporaryDirectory() as folder:
            filename = Path(folder) / "export.xlsx"
            exported = export_spreadsheet(filename, _Repository())

            self.assertEqual(exported, 1)
            workbook = load_workbook(filename, data_only=False)
            worksheet = workbook["Drills"]
            self.assertEqual(
                tuple(cell.value for cell in worksheet[1]), EXPORT_HEADERS
            )
            self.assertEqual(worksheet.max_row, 2)
            self.assertEqual(worksheet["B2"].value, "Receiving & Passing")
            self.assertEqual(worksheet["C2"].value, "First Touch")
            self.assertEqual(worksheet["E2"].value, "'=unsafe formula text")
            self.assertEqual(worksheet["M2"].value, "Balls | Cones")
            self.assertEqual(worksheet.freeze_panes, "A2")
            self.assertEqual(worksheet.auto_filter.ref, "A1:Q2")

    def test_exported_drills_can_be_imported_with_full_details(self):
        with tempfile.TemporaryDirectory() as folder:
            filename = Path(folder) / "export.xlsx"
            export_spreadsheet(filename, _Repository())
            destination = _Repository([])

            report = import_spreadsheet(filename, destination)

            self.assertEqual(report.imported, 1)
            self.assertEqual(report.duplicates, [])
            self.assertEqual(report.errors, [])
            imported = destination.drills[0]
            self.assertEqual(imported.name, "Passing Pattern")
            self.assertEqual(imported.development_block_id, 2)
            self.assertIsNone(imported.technical_focus_id)
            self.assertEqual(imported.coaching_focus, "First Touch")
            self.assertTrue(imported.use_execution_details)
            self.assertEqual((imported.sets, imported.reps), (3, 4))
            self.assertEqual((imported.work_seconds, imported.rest_seconds), (45, 20))
            self.assertEqual(imported.equipment, ["Balls", "Cones"])
            self.assertEqual(imported.coaching_points, ["Scan", "Open body"])

    def test_export_uses_current_database_block_ids(self):
        drills = [
            Drill(
                id=9,
                name="Modern Block Drill",
                development_block_id=610,
            )
        ]
        with tempfile.TemporaryDirectory() as folder:
            filename = Path(folder) / "export.xlsx"

            export_spreadsheet(filename, _Repository(drills), _BlockRepository())

            worksheet = load_workbook(filename)["Drills"]
            self.assertEqual(worksheet["B2"].value, "Receiving & Passing")

    def test_legacy_active_drills_sheet_can_still_be_imported(self):
        with tempfile.TemporaryDirectory() as folder:
            filename = Path(folder) / "legacy-export.xlsx"
            export_spreadsheet(filename, _Repository())
            workbook = load_workbook(filename)
            workbook["Drills"].title = "Active Drills"
            workbook.save(filename)
            destination = _Repository([])

            report = import_spreadsheet(filename, destination)

            self.assertEqual(report.imported, 1)
            self.assertEqual(report.errors, [])

    def test_created_blank_template_uses_complete_schema(self):
        with tempfile.TemporaryDirectory() as folder:
            filename = Path(folder) / "template.xlsx"
            create_template(filename)
            workbook = load_workbook(filename)

            self.assertEqual(
                tuple(cell.value for cell in workbook["Drills"][1]),
                TEMPLATE_HEADERS,
            )
            self.assertEqual(workbook["Drills"].max_row, 1)
            self.assertEqual(workbook["Drills"].freeze_panes, "A2")
            self.assertEqual(workbook["Drills"].auto_filter.ref, "A1:Q1")

    def test_missing_drill_sheet_returns_a_useful_error(self):
        with tempfile.TemporaryDirectory() as folder:
            filename = Path(folder) / "wrong-sheet.xlsx"
            workbook = Workbook()
            workbook.active.title = "Other"
            workbook.save(filename)

            report = import_spreadsheet(filename, _Repository([]))

            self.assertEqual(report.imported, 0)
            self.assertEqual(
                report.errors,
                ['The spreadsheet must contain a "Drills" or "Active Drills" sheet.'],
            )


if __name__ == "__main__":
    unittest.main()
